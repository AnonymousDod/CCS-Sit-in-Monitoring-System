from flask import Flask, request, render_template, redirect, url_for, flash, session, jsonify, Response, make_response, send_file
import sqlite3
import os
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import io
import csv
from io import StringIO
from functools import wraps
import random
import string
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

app = Flask(__name__)
app.secret_key = 'your_secret_key'

UPLOAD_FOLDER = "static/uploads/"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

# Admin credentials (in a real application, these should be stored securely in a database)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"  # In production, use a secure password

# Ensure uploads directory exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Function to initialize database with admin table
def init_db():
    try:
        conn = sqlite3.connect('users.db')
        c = conn.cursor()
        
        # Create users table
        c.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                firstname TEXT NOT NULL,
                midname TEXT,
                lastname TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                course TEXT NOT NULL,
                yearlevel TEXT NOT NULL,
                is_admin INTEGER DEFAULT 0,
                profile_pic TEXT,
                remaining_sessions INTEGER DEFAULT 0,
                points INTEGER DEFAULT 0
            )
        ''')
        
        # Create sessions table
        c.execute('''
            CREATE TABLE IF NOT EXISTS sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                start_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                end_time TIMESTAMP,
                status TEXT DEFAULT 'active',
                purpose TEXT,
                lab_unit TEXT,
                duration INTEGER,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        # Create sit_in_history table
        c.execute('''
            CREATE TABLE IF NOT EXISTS sit_in_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                date DATE NOT NULL,
                time_in TIMESTAMP NOT NULL,
                time_out TIMESTAMP,
                status TEXT DEFAULT 'active',
                purpose TEXT,
                lab_unit TEXT,
                duration INTEGER,
                allocated_duration INTEGER,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        # Create reservations table
        c.execute('''
            CREATE TABLE IF NOT EXISTS reservations (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                time TEXT NOT NULL,
                purpose TEXT NOT NULL,
                lab_unit TEXT NOT NULL,
                status TEXT DEFAULT 'pending',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')
        
        # Create announcements table
        c.execute('''
            CREATE TABLE IF NOT EXISTS announcements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                content TEXT NOT NULL,
                priority TEXT DEFAULT 'normal',
                created_by TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create feedback table
        c.execute('''
            CREATE TABLE IF NOT EXISTS feedback (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                rating INTEGER NOT NULL,
                comments TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users (id)
            )
        ''')

        # Create lab_resources table
        c.execute('''
            CREATE TABLE IF NOT EXISTS lab_resources (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                status TEXT DEFAULT 'enabled',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Insert default lab resources if they don't exist
        c.execute("SELECT COUNT(*) FROM lab_resources")
        if c.fetchone()[0] == 0:
            default_resources = [
                ('Lab Unit 1',),
                ('Lab Unit 2',),
                ('Lab Unit 3',),
                ('Lab Unit 4',)
            ]
            c.executemany("INSERT INTO lab_resources (name) VALUES (?)", default_resources)
        
        # Check if admin user exists, if not create it
        c.execute("SELECT * FROM users WHERE username = ?", (ADMIN_USERNAME,))
        if not c.fetchone():
            admin_password_hash = generate_password_hash(ADMIN_PASSWORD, method="pbkdf2:sha256")
            c.execute("""
                INSERT INTO users (username, password, firstname, lastname, midname, email, course, yearlevel, is_admin, profile_pic)
                VALUES (?, ?, 'Admin', 'User', '', 'admin@example.com', 'N/A', 'N/A', 1, 'default.png')
            """, (ADMIN_USERNAME, admin_password_hash))
            print("Admin user created successfully!")
        
        conn.commit()
        print("Database initialized successfully!")
        
        # Update existing users with remaining sessions
        try:
            c.execute("ALTER TABLE users ADD COLUMN remaining_sessions INTEGER DEFAULT 0")
            conn.commit()
            print("Added remaining_sessions column to users table")
        except sqlite3.OperationalError:
            # Column already exists, ignore the error
            pass
            
        # Add points column to users table
        try:
            c.execute("ALTER TABLE users ADD COLUMN points INTEGER DEFAULT 0")
            conn.commit()
            print("Added points column to users table")
        except sqlite3.OperationalError:
            # Column already exists, ignore the error
            pass
            
    except Exception as e:
        print(f"Error initializing database: {e}")
        raise e
    finally:
        if 'conn' in locals():
            conn.close()

# Initialize database when app starts
with app.app_context():
    init_db()

# Function to fetch user details
def get_user(username):
    con = sqlite3.connect("users.db")
    con.row_factory = sqlite3.Row  
    cur = con.cursor()
    
    cur.execute("SELECT * FROM users WHERE username = ?", (username,))
    user = cur.fetchone()
    con.close()
    
    return dict(user) if user else None  

# Function to update session after profile changes
def update_session(username):
    user = get_user(username)
    if user:
        session["user"] = {
            "id": user["id"],
            "firstname": user["firstname"],
            "lastname": user["lastname"],
            "email": user["email"],
            "course": user["course"],
            "yearlevel": user["yearlevel"],
            "profile_pic": user["profile_pic"] if user["profile_pic"] else "default.png",
            "is_admin": user["is_admin"]
        }
        if user["is_admin"]:
            session["admin"] = True

# Admin authentication decorator
def admin_required(f):
    def decorated_function(*args, **kwargs):
        print("==== ADMIN REQUIRED DECORATOR ====")
        print("Current session:", session)
        print("'admin' in session:", 'admin' in session)
        if 'admin' in session:
            print("session['admin'] value:", session['admin'])
        if 'user' in session:
            print("User in session:", session['user'])
            if 'is_admin' in session['user']:
                print("user['is_admin'] value:", session['user']['is_admin'])
        
        if "admin" not in session or not session["admin"]:
            print("Admin authentication failed, redirecting to login")
            flash("Please login as admin first.", "error")
            return redirect(url_for("login_page"))
        
        print("Admin authentication successful")
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Admin routes
@app.route("/admin/dashboard")
@admin_required
def admin_dashboard():
    # Get statistics
    conn = sqlite3.connect('users.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Get active users count
    cursor.execute("SELECT COUNT(*) FROM users WHERE is_admin = 0")
    active_users = cursor.fetchone()[0]
    
    # Get active sessions count
    cursor.execute("SELECT COUNT(*) FROM sit_in_history WHERE time_out IS NULL")
    active_sessions = cursor.fetchone()[0]
    
    # Get total records count
    cursor.execute("SELECT COUNT(*) FROM sit_in_history")
    total_records = cursor.fetchone()[0]
    
    # Get today's sessions count
    today = datetime.now().strftime('%Y-%m-%d')
    cursor.execute("SELECT COUNT(*) FROM sit_in_history WHERE date(time_in) = ?", (today,))
    today_sessions = cursor.fetchone()[0]
    
    # Get total hours for today
    cursor.execute("""
        SELECT SUM(
            CASE
                WHEN time_out IS NULL THEN (strftime('%s', 'now') - strftime('%s', time_in)) / 60
                ELSE (strftime('%s', time_out) - strftime('%s', time_in)) / 60
            END
        ) AS total_minutes
        FROM sit_in_history
        WHERE date(time_in) = ?
    """, (today,))
    result = cursor.fetchone()
    total_minutes_today = result[0] if result[0] is not None else 0
    total_hours_today = round(total_minutes_today / 60, 1)
    
    # Get sit-in purpose distribution for the chart
    cursor.execute("""
        SELECT 
            CASE 
                WHEN purpose IS NULL OR purpose = '' THEN 'Other'
                ELSE purpose 
            END as purpose, 
            COUNT(*) as count 
        FROM sit_in_history 
        GROUP BY purpose
        ORDER BY count DESC
    """)
    purpose_stats = [dict(row) for row in cursor.fetchall()]
    
    # Still collect year level statistics for potential future use
    cursor.execute("""
        SELECT yearlevel, COUNT(*) as count 
        FROM users 
        WHERE is_admin = 0 
        GROUP BY yearlevel
        ORDER BY 
            CASE yearlevel
                WHEN '1st Year' THEN 1
                WHEN '2nd Year' THEN 2
                WHEN '3rd Year' THEN 3
                WHEN '4th Year' THEN 4
                ELSE 5
            END
    """)
    year_level_stats = [dict(row) for row in cursor.fetchall()]
    
    # Get announcements
    cursor.execute("""
        SELECT id, title, content, priority, created_by, created_at
        FROM announcements
        ORDER BY created_at DESC
    """)
    announcements = cursor.fetchall()
    
    conn.close()
    
    stats = {
        'active_users': active_users,
        'active_sessions': active_sessions,
        'total_records': total_records,
        'today_sessions': today_sessions,
        'total_hours_today': total_hours_today
    }
    
    return render_template(
        "admin_dashboard.html", 
        stats=stats, 
        purpose_stats=purpose_stats,
        year_level_stats=year_level_stats,
        announcements=announcements
    )

@app.route("/admin/users")
@admin_required
def admin_users():
    con = sqlite3.connect("users.db")
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute("SELECT * FROM users ORDER BY id DESC")
    users = cur.fetchall()
    
    # Get course distribution
    cur.execute("""
        SELECT course, COUNT(*) as count 
        FROM users 
        WHERE is_admin = 0
        GROUP BY course
        ORDER BY count DESC
    """)
    course_stats = [dict(zip([column[0] for column in cur.description], row)) for row in cur.fetchall()]
    
    # Get year level distribution
    cur.execute("""
        SELECT yearlevel, COUNT(*) as count 
        FROM users 
        WHERE is_admin = 0
        GROUP BY yearlevel
        ORDER BY yearlevel
    """)
    year_level_stats = [dict(zip([column[0] for column in cur.description], row)) for row in cur.fetchall()]
    
    # Get recent users
    cur.execute("""
        SELECT id, username, firstname, lastname, email, course, yearlevel
        FROM users
        WHERE is_admin = 0
        ORDER BY id DESC
        LIMIT 5
    """)
    recent_users = [dict(zip([column[0] for column in cur.description], row)) for row in cur.fetchall()]
    
    con.close()
    return render_template("admin_users.html", users=users, course_stats=course_stats, year_level_stats=year_level_stats, recent_users=recent_users)

@app.route("/admin/sessions")
@admin_required
def admin_sessions():
    # Get database connection
    db = get_db()
    cursor = db.cursor()
    
    # Get session statistics
    cursor.execute("""
        SELECT 
            COUNT(CASE WHEN time_out IS NULL THEN 1 END) as active_sessions,
            COUNT(*) as total_records,
            COUNT(CASE WHEN DATE(time_in) = DATE('now', 'localtime') THEN 1 END) as today_sessions,
            SUM(CASE 
                WHEN DATE(time_in) = DATE('now', 'localtime') THEN 
                    CASE 
                        WHEN time_out IS NOT NULL THEN 
                            (julianday(time_out) - julianday(time_in)) * 24
                        ELSE 
                            (julianday('now', 'localtime') - julianday(time_in)) * 24
                    END
                ELSE 0 
            END) as total_hours_today
        FROM sit_in_history
    """)
    stats_data = cursor.fetchone()
    
    # Format the stats data into a dictionary
    stats = {}
    if stats_data:
        stats = dict(zip(['active_sessions', 'total_records', 'today_sessions', 'total_hours_today'], stats_data))
        
        # Format the total hours
        if stats['total_hours_today'] is not None:
            stats['total_hours_today'] = f"{stats['total_hours_today']:.1f}"
        else:
            stats['total_hours_today'] = "0.0"
    else:
        # Default values if no data found
        stats = {
            'active_sessions': 0,
            'total_records': 0,
            'today_sessions': 0,
            'total_hours_today': "0.0"
        }
    
    # Render the template with statistics
    return render_template("admin_sessions.html", stats=stats)

@app.route("/admin/logout")
def admin_logout():
    # Clear all session data
    session.clear()
    print("Admin logged out, session cleared:", session)
    flash("Admin logged out successfully!", "success")
    return redirect(url_for("login_page"))

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == "POST":
        idno = request.form.get("idno")
        firstname = request.form.get("firstname")
        lastname = request.form.get("lastname")
        midname = request.form.get("midname")
        course = request.form.get("course")
        yearlevel = request.form.get("yearlevel")
        email = request.form.get("email")
        password = request.form.get("password")

        # Ensure all fields are filled
        if not (idno and firstname and lastname and midname and course and yearlevel and email and password):
            flash("All fields are required!", "error")
            return redirect(url_for("signup"))
        
        # Check if ID number already exists
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        cur.execute("SELECT * FROM users WHERE username = ?", (idno,))
        existing_user = cur.fetchone()

        if existing_user:
            flash("ID Number already registered. Please use a different ID number.", "error")
            con.close()
            return redirect(url_for("signup"))

        # Hash the password
        hashed_password = generate_password_hash(password, method="pbkdf2:sha256")

        # Map course values to actual course names
        course_map = {
            "1": "Information Technology",
            "2": "Computer Engineering",
            "3": "Criminology",
            "4": "Customs Administration"
        }
        course_name = course_map.get(course, "Unknown Course")

        # Map year level values to actual year levels
        year_map = {
            "1": "1st Year",
            "2": "2nd Year",
            "3": "3rd Year",
            "4": "4th Year"
        }
        year_name = year_map.get(yearlevel, "Unknown Year")

        # Insert user into the database using ID number as username
        cur.execute("""
            INSERT INTO users (firstname, lastname, midname, email, username, password, course, yearlevel, profile_pic, remaining_sessions)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 30)
        """, (firstname, lastname, midname, email, idno, hashed_password, course_name, year_name, "default.png"))

        con.commit()
        con.close()

        flash("Registration successful! You can now log in using your ID number.", "success")
        return redirect(url_for("login_page"))

    return render_template("signup.html")

# Function to check login credentials
def check_login(username, password, is_admin=False):
    user = get_user(username)
    if user and check_password_hash(user["password"], password):
        if is_admin:
            if user["is_admin"]:
                update_session(username)
                return True
            return False
        elif user["is_admin"]:
            # If user is admin but didn't check the box, prevent regular login
            return False
        update_session(username)
        return True
    return False

@app.route("/", methods=["GET", "POST"])
@app.route("/login", methods=["GET", "POST"])
def login_page():
    if request.method == "POST":
        username = request.form.get("un")
        password = request.form.get("pwd")
        is_admin = request.form.get("is_admin") == "on"
        
        print(f"Login attempt: username={username}, is_admin={is_admin}")

        if not username or not password:
            flash('Please fill out both fields.', 'error')
            return redirect(url_for('login_page'))
        
        # Get user from database
        user = get_user(username)
        if not user:
            flash('Invalid ID number or password.', 'error')
            return redirect(url_for('login_page'))
            
        # Verify password
        if not check_password_hash(user["password"], password):
            flash('Invalid ID number or password.', 'error')
            return redirect(url_for('login_page'))
            
        # Set up user session
        session["user"] = {
            "id": user["id"],
            "firstname": user["firstname"],
            "lastname": user["lastname"],
            "email": user["email"],
            "course": user["course"],
            "yearlevel": user["yearlevel"],
            "profile_pic": user["profile_pic"] if user["profile_pic"] else "default.png",
            "is_admin": bool(user["is_admin"])
        }
        
        # Admin login
        if is_admin:
            if user["is_admin"]:
                session["admin"] = True
                print(f"Admin login successful: {session}")
                flash('Admin login successful!', 'success')
                return redirect(url_for("admin_dashboard"))
            else:
                session.pop("user", None)
                flash('You do not have admin privileges.', 'error')
                return redirect(url_for('login_page'))
        else:
            # Regular user login
            if user["is_admin"]:
                session.pop("user", None)
                flash('Please check "Login as Administrator" to access the admin dashboard.', 'error')
                return redirect(url_for('login_page'))
                
            return redirect(url_for("home"))
    
    return render_template("login.html")

@app.route("/home")
@app.route("/home/<show_section>")
def home(show_section=None):
    if "user" not in session:
        flash("Please log in first.", "error")
        return redirect(url_for("login_page"))
    
    # Check if user is an admin
    if session["user"].get("is_admin"):
        flash("Administrators should use the admin dashboard.", "error")
        return redirect(url_for("admin_dashboard"))
    
    # Get announcements
    con = sqlite3.connect("users.db")
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    cur.execute("""
        SELECT a.*, u.username as created_by
        FROM announcements a
        JOIN users u ON a.created_by = u.id
        ORDER BY a.created_at DESC
        LIMIT 5
    """)
    announcements = []
    for row in cur.fetchall():
        announcement = dict(row)
        # Convert created_at string to datetime object
        announcement['created_at'] = datetime.strptime(announcement['created_at'], '%Y-%m-%d %H:%M:%S')
        announcements.append(announcement)
    
    # Get user's history data (always load this for improved performance)
    user_id = session["user"]["id"]
    
    # Get user's session history
    cur.execute("""
            SELECT id, date, time_in, time_out, status, duration, allocated_duration, purpose, lab_unit 
            FROM sit_in_history
            WHERE user_id = ?
            ORDER BY time_in DESC
    """, (user_id,))
    history = cur.fetchall()
    
    # Get statistics
    cur.execute("""
        SELECT 
            COUNT(*) as total_sessions,
            COUNT(CASE WHEN status = 'completed' THEN 1 END) as completed_sessions,
            SUM(CASE WHEN duration IS NOT NULL THEN duration ELSE 0 END) as total_duration
        FROM sit_in_history
        WHERE user_id = ?
    """, (user_id,))
    stats = cur.fetchone()
    
    con.close()
    
    # Format history data
    formatted_history = []
    for entry in history:
        formatted_entry = {
            'id': entry[0],
            'date': entry[1],
            'time_in': entry[2],
            'time_out': entry[3] if entry[3] else None,
            'status': entry[4],
            'duration': entry[5],
            'allocated_duration': entry[6],
            'purpose': entry[7],
            'lab_unit': entry[8]
        }
        formatted_history.append(formatted_entry)
    
    # Format statistics
    statistics = {
        'total_sessions': stats[0],
        'completed_sessions': stats[1],
        'total_duration': stats[2]
    }
    
    return render_template(
        "homepage.html", 
        user=session["user"], 
        announcements=announcements,
        history=formatted_history,
        stats=statistics,
        show_section=show_section
    )

# Route for updating profile details
@app.route("/update_profile", methods=["POST"])
def update_profile():
    if "user" not in session:
        flash("Please log in first.", "error")
        return redirect(url_for("login_page"))

    user_id = session["user"]["id"]
    firstname = request.form.get("firstname")
    lastname = request.form.get("lastname")
    email = request.form.get("email")
    course = request.form.get("course")
    yearlevel = request.form.get("yearlevel")

    if not (firstname and lastname and email and course and yearlevel):
        flash("All fields are required!", "error")
        return redirect(url_for("home"))

    # Handle profile picture upload if provided
    profile_pic = session["user"].get("profile_pic", "default.png")  # Keep existing picture by default
    
    if "profile_pic" in request.files:
        file = request.files["profile_pic"]
        if file and file.filename:
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            file.save(filepath)
            profile_pic = filename

    # Update database with all information including profile picture
    con = sqlite3.connect("users.db")
    cur = con.cursor()
    cur.execute("""
        UPDATE users 
        SET firstname = ?, lastname = ?, email = ?, course = ?, yearlevel = ?, profile_pic = ?
        WHERE id = ?
    """, (firstname, lastname, email, course, yearlevel, profile_pic, user_id))
    con.commit()
    con.close()

    # Update session data
    session["user"] = {
        "id": user_id,
        "firstname": firstname,
        "lastname": lastname,
        "email": email,
        "course": course,
        "yearlevel": yearlevel,
        "profile_pic": profile_pic
    }

    flash("Profile updated successfully!", "success")
    return redirect(url_for("home"))

@app.route("/profile")
def profile():
    if "user" not in session:
        flash("Please log in first.", "error")
        return redirect(url_for("login_page"))

    user_id = session["user"]["id"]
    
    con = sqlite3.connect("users.db")
    cur = con.cursor()
    cur.execute("SELECT firstname, lastname, email, course FROM users WHERE id = ?", (user_id,))
    user = cur.fetchone()
    con.close()

    if user:
        session["user"]["firstname"] = user[0]
        session["user"]["lastname"] = user[1]
        session["user"]["email"] = user[2]
        session["user"]["course"] = user[3]  # Ensure course is stored in session
    else:
        flash("User not found.", "error")
        return redirect(url_for("home"))

    return redirect(url_for("profile"))

@app.route('/info')
def info():
    return render_template('info.html')

@app.route('/edit')
def edit():
    return render_template('edit.html')

def get_db():
    con = sqlite3.connect("users.db")
    con.row_factory = sqlite3.Row
    return con

@app.route('/announcements')
def announcements():
    if "user" not in session:
        flash("Please log in first.", "error")
        return redirect(url_for("login_page"))
    
    # Get announcements
    con = get_db()
    cur = con.cursor()
    cur.execute("""
        SELECT a.*, u.username as created_by
        FROM announcements a
        JOIN users u ON a.created_by = u.id
        ORDER BY a.created_at DESC
    """)
    announcements = [dict(row) for row in cur.fetchall()]
    con.close()
    
    return render_template("announcements.html", announcements=announcements)

@app.route('/remaining_sessions')
def remaining_sessions():
    if "user" not in session:
        flash("Please log in first.", "error")
        return redirect(url_for("login_page"))
    
    user_id = session["user"]["id"]
    
    con = sqlite3.connect("users.db")
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    
    # Get active sessions for the user
    cur.execute("""
        SELECT s.*, u.firstname, u.lastname
        FROM sessions s
        JOIN users u ON s.user_id = u.id
        WHERE s.user_id = ? AND s.status = 'active'
        ORDER BY s.start_time DESC
    """, (user_id,))
    active_sessions = [dict(row) for row in cur.fetchall()]
    
    # Get completed sessions for the last 7 days
    cur.execute("""
        SELECT s.*, u.firstname, u.lastname
        FROM sessions s
        JOIN users u ON s.user_id = u.id
        WHERE s.user_id = ? 
        AND s.status = 'completed'
        AND s.start_time >= datetime('now', '-7 days')
        ORDER BY s.start_time DESC
    """, (user_id,))
    recent_sessions = [dict(row) for row in cur.fetchall()]
    
    con.close()
    
    return render_template('remaining_sessions.html', 
                         active_sessions=active_sessions,
                         recent_sessions=recent_sessions)

@app.route('/sit_in_rors')
def sit_in_rors():
    return render_template('sit_in_rors.html')

@app.route('/sit_in_history')
def sit_in_history():
    if "user" not in session:
        return redirect(url_for('login_page'))
    
    # Redirect to the homepage with the history section shown
    return redirect(url_for('home', show_section='history'))

@app.route('/register_user', methods=['GET', 'POST'])
def register_user():
    return render_template('signup.html')

@app.route("/logout")
def logout():
    if "admin" in session:
        session.pop("admin", None)
        flash("Admin logged out successfully!", "success")
    else:
        session.clear()
        flash("You have been logged out.", "success")
    return redirect(url_for("login_page"))

# Remove the register_admin route since we're using a fixed admin account
@app.route("/register_admin", methods=["GET", "POST"])
def register_admin():
    flash("Admin registration is not available.", "error")
    return redirect(url_for("login_page"))

# Add new route for starting a session
@app.route('/start_session', methods=['POST'])
def start_session():
    if "user" not in session:
        return jsonify({"error": "Please log in first."}), 401
    
    user_id = session["user"]["id"]
    current_time = datetime.now()
    
    try:
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        
        # Check if user already has an active session
        cur.execute("""
            SELECT id FROM sessions 
            WHERE user_id = ? AND status = 'active'
        """, (user_id,))
        
        if cur.fetchone():
            con.close()
            return jsonify({"error": "You already have an active session"}), 400
        
        # Create new session
        cur.execute("""
            INSERT INTO sessions (user_id, start_time, status)
            VALUES (?, ?, 'active')
        """, (user_id, current_time))
        
        session_id = cur.lastrowid
        
        con.commit()
        con.close()
        
        return jsonify({"success": True, "message": "Session started successfully", "session_id": session_id})
    except Exception as e:
        if con:
            con.close()
        return jsonify({"error": str(e)}), 500

@app.route('/end_session', methods=['POST'])
def end_session():
    if "user" not in session:
        return jsonify({"error": "Please log in first."}), 401
    
    user_id = session["user"]["id"]
    current_time = datetime.now()
    
    try:
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        
        # Get active session for the user
        cur.execute("""
            SELECT id, start_time 
            FROM sessions 
            WHERE user_id = ? AND status = 'active'
            ORDER BY start_time DESC LIMIT 1
        """, (user_id,))
        active_session = cur.fetchone()

        if not active_session:
            return jsonify({'error': 'No active session found'}), 404

        # Calculate duration in minutes
        start_time = datetime.strptime(active_session[1], '%Y-%m-%d %H:%M:%S.%f')
        duration = int((current_time - start_time).total_seconds() / 60)

        # Update session
        cur.execute("""
            UPDATE sessions 
            SET end_time = ?, status = 'completed', duration = ?
            WHERE id = ?
        """, (current_time, duration, active_session[0]))

        con.commit()
        con.close()
        return jsonify({'success': True})
    except Exception as e:
        if con:
            con.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/search_student/<student_id>")
@admin_required
def search_student(student_id):
    con = sqlite3.connect("users.db")
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    
    cur.execute("""
        SELECT id, username, firstname, lastname, course, yearlevel
        FROM users
        WHERE username = ? AND is_admin = 0
    """, (student_id,))
    
    student = cur.fetchone()
    con.close()
    
    if student:
        return jsonify({
            "found": True,
            "id": student["id"],
            "firstname": student["firstname"],
            "lastname": student["lastname"],
            "course": student["course"],
            "yearlevel": student["yearlevel"]
        })
    
    return jsonify({"found": False})

@app.route("/api/start_sitin", methods=["POST"])
@admin_required
def start_sitin():
    data = request.json
    student_id = data.get('student_id')
    purpose = data.get('purpose')
    lab_unit = data.get('lab_unit')
    duration = data.get('duration', 60)  # Default to 60 minutes if not specified

    try:
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        
        # First, ensure the sit_in_history table exists
        cur.execute('''
        CREATE TABLE IF NOT EXISTS sit_in_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            date DATE NOT NULL,
            time_in TIMESTAMP NOT NULL,
            time_out TIMESTAMP,
            status TEXT DEFAULT 'active',
            purpose TEXT,
            lab_unit TEXT,
            duration INTEGER,
            allocated_duration INTEGER,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
        ''')
        con.commit()
        
        current_time = datetime.now()

        # Get user_id and remaining_sessions from student_id
        cur.execute("SELECT id, remaining_sessions, firstname, lastname FROM users WHERE username = ?", (student_id,))
        user_data = cur.fetchone()
        
        if not user_data:
            return jsonify({"success": False, "error": "Student not found"})
        
        user_id, remaining_sessions, firstname, lastname = user_data

        # Check if user has remaining sessions
        if remaining_sessions <= 0:
            return jsonify({"success": False, "error": "No remaining sessions available"})
            
        # Check if user already has an active sit-in session
        cur.execute("""
            SELECT id FROM sit_in_history 
            WHERE user_id = ? AND status = 'active'
        """, (user_id,))
        
        if cur.fetchone():
            return jsonify({"success": False, "error": "Student already has an active sit-in session"})

        # Start the sit-in session
        cur.execute("""
            INSERT INTO sit_in_history (user_id, date, time_in, status, purpose, lab_unit, allocated_duration)
            VALUES (?, ?, ?, 'active', ?, ?, ?)
        """, (user_id, current_time.date(), current_time, purpose, lab_unit, duration))
        
        session_id = cur.lastrowid
        
        # Deduct one session from remaining_sessions
        cur.execute("""
            UPDATE users 
            SET remaining_sessions = remaining_sessions - 1
            WHERE id = ?
        """, (user_id,))
        
        # Get updated remaining sessions
        new_remaining_sessions = remaining_sessions - 1
        
        con.commit()
        con.close()
        
        return jsonify({
            "success": True,
            "message": "Sit-in session started successfully",
            "session_id": session_id,
            "student_name": f"{firstname} {lastname}",
            "remaining_sessions": new_remaining_sessions
        })
    except Exception as e:
        if 'con' in locals():
            con.close()
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/end_sitin/<int:session_id>", methods=["POST"])
@admin_required
def end_sitin(session_id):
    try:
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        current_time = datetime.now()
        
        # Update sit-in history directly using session_id
        cur.execute("""
            UPDATE sit_in_history 
            SET time_out = ?, 
                status = 'completed',
                duration = CAST((julianday(?) - julianday(time_in)) * 24 * 3600 AS INTEGER)
            WHERE id = ? AND status = 'active'
        """, (current_time, current_time, session_id))
        
        con.commit()
        return jsonify({"success": True})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        con.close()

@app.route("/api/delete_sitin/<int:session_id>", methods=["POST"])
@admin_required
def delete_sitin(session_id):
    try:
        # Get session purpose before deletion (for logging)
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        
        # Get the purpose of the session to be deleted
        cur.execute("SELECT purpose FROM sit_in_history WHERE id = ?", (session_id,))
        session_data = cur.fetchone()
        
        if not session_data:
            return jsonify({"success": False, "error": "Session not found"}), 404
            
        # Delete the session
        cur.execute("DELETE FROM sit_in_history WHERE id = ?", (session_id,))
        
        if cur.rowcount == 0:
            return jsonify({"success": False, "error": "No session was deleted"}), 404
            
        con.commit()
        return jsonify({"success": True, "message": "Session deleted successfully"})
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        if 'con' in locals():
            con.close()

@app.route('/make_reservation', methods=['POST'])
def make_reservation():
    if "user" not in session:
        return jsonify({"error": "Please log in first."}), 401
    
    data = request.get_json()
    user_id = session["user"]["id"]
    
    try:
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        
        # Insert new reservation
        cur.execute("""
            INSERT INTO reservations (user_id, date, time, purpose, lab_unit, status)
            VALUES (?, ?, ?, ?, ?, 'pending')
        """, (user_id, data['date'], data['time'], data['purpose'], data['lab_unit']))
        
        con.commit()
        con.close()
        
        return jsonify({"success": True, "message": "Reservation submitted successfully"})
    except Exception as e:
        if con:
            con.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/get_reservations")
@admin_required
def get_reservations():
    try:
        con = sqlite3.connect("users.db")
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        
        # Debug: Print all reservations regardless of status
        cur.execute("""
            SELECT r.*, u.firstname, u.lastname, u.course, u.yearlevel
            FROM reservations r
            JOIN users u ON r.user_id = u.id
            ORDER BY r.created_at DESC
        """)
        all_reservations = [dict(row) for row in cur.fetchall()]
        print("All reservations:", all_reservations)
        
        # Get pending reservations
        cur.execute("""
            SELECT r.*, u.firstname, u.lastname, u.course, u.yearlevel
            FROM reservations r
            JOIN users u ON r.user_id = u.id
            WHERE r.status = 'pending'
            ORDER BY r.created_at DESC
        """)
        
        pending_reservations = [dict(row) for row in cur.fetchall()]
        print("Pending reservations:", pending_reservations)
        
        con.close()
        
        return jsonify({
            "reservations": pending_reservations,
            "debug_all_reservations": all_reservations  # This will help us debug
        })
        
    except Exception as e:
        print("Error in get_reservations:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route("/admin/reservations")
@admin_required
def admin_reservations():
    return render_template("admin_reservations.html")

@app.route("/api/get_pending_reservations")
@admin_required
def get_pending_reservations():
    try:
        con = get_db()
        cur = con.cursor()
        
        # Debug: Check if reservations table exists
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='reservations'")
        if not cur.fetchone():
            return jsonify({"success": False, "error": "Reservations table does not exist"})
        
        # Debug: Get all reservations regardless of status
        cur.execute("""
            SELECT r.*, u.firstname, u.lastname, u.course, u.yearlevel
            FROM reservations r
            JOIN users u ON r.user_id = u.id
            ORDER BY r.date, r.time
        """)
        all_reservations = [dict(row) for row in cur.fetchall()]
        
        # Get pending reservations
        cur.execute("""
            SELECT r.*, u.firstname, u.lastname, u.course, u.yearlevel
            FROM reservations r
            JOIN users u ON r.user_id = u.id
            WHERE r.status = 'pending'
            ORDER BY r.date, r.time
        """)
        pending_reservations = [dict(row) for row in cur.fetchall()]
        
        con.close()
        
        return jsonify({
            "success": True, 
            "reservations": pending_reservations,
            "debug": {
                "all_count": len(all_reservations),
                "pending_count": len(pending_reservations),
                "all_reservations": all_reservations[:5]  # Just send first 5 for debugging
            }
        })
    except Exception as e:
        print("Error in get_pending_reservations:", str(e))
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/approve_reservation/<int:reservation_id>", methods=["POST"])
@admin_required
def approve_reservation(reservation_id):
    con = get_db()
    cur = con.cursor()
    try:
        cur.execute("UPDATE reservations SET status = 'approved' WHERE id = ?", (reservation_id,))
        con.commit()
        return jsonify({"success": True, "message": "Reservation approved successfully"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        con.close()

@app.route("/api/reject_reservation/<int:reservation_id>", methods=["POST"])
@admin_required
def reject_reservation(reservation_id):
    con = get_db()
    cur = con.cursor()
    try:
        cur.execute("UPDATE reservations SET status = 'rejected' WHERE id = ?", (reservation_id,))
        con.commit()
        return jsonify({"success": True, "message": "Reservation rejected successfully"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        con.close()

@app.route("/api/delete_user/<int:user_id>", methods=["POST"])
@admin_required
def delete_user(user_id):
    try:
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        
        # Check if user exists and is not an admin
        cur.execute("SELECT is_admin FROM users WHERE id = ?", (user_id,))
        user = cur.fetchone()
        if not user:
            return jsonify({"success": False, "error": "User not found"})
        if user[0]:
            return jsonify({"success": False, "error": "Cannot delete admin users"})
        
        # Delete user's sessions
        cur.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
        # Delete user's sit-in history
        cur.execute("DELETE FROM sit_in_history WHERE user_id = ?", (user_id,))
        # Delete user's reservations
        cur.execute("DELETE FROM reservations WHERE user_id = ?", (user_id,))
        # Delete the user
        cur.execute("DELETE FROM users WHERE id = ?", (user_id,))
        
        con.commit()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        con.close()

@app.route("/api/update_user/<int:user_id>", methods=["POST"])
@admin_required
def update_user(user_id):
    try:
        data = request.json
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        
        # Check if user exists
        cur.execute("SELECT is_admin FROM users WHERE id = ?", (user_id,))
        if not cur.fetchone():
            return jsonify({"success": False, "error": "User not found"})
        
        # Update user details
        cur.execute("""
            UPDATE users 
            SET firstname = ?, lastname = ?, email = ?, 
                course = ?, yearlevel = ?
            WHERE id = ?
        """, (data['firstname'], data['lastname'], data['email'],
              data['course'], data['yearlevel'], user_id))
        
        con.commit()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        con.close()

@app.route("/api/get_user_details/<int:user_id>")
@admin_required
def get_user_details(user_id):
    try:
        con = sqlite3.connect("users.db")
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        
        # Get user details
        cur.execute("""
            SELECT id, username, firstname, lastname, email, 
                   course, yearlevel, is_admin
            FROM users WHERE id = ?
        """, (user_id,))
        user = cur.fetchone()
        
        if not user:
            return jsonify({"success": False, "error": "User not found"})
        
        # Get user's session statistics
        cur.execute("""
            SELECT 
                COUNT(*) as total_sessions,
                SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed_sessions,
                SUM(CASE WHEN status = 'active' THEN 1 ELSE 0 END) as active_sessions,
                SUM(duration) as total_duration
            FROM sessions 
            WHERE user_id = ?
        """)
        stats = cur.fetchone()
        
        return jsonify({
            "success": True,
            "user": dict(user),
            "stats": dict(stats)
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        con.close()

@app.route("/api/get_session_stats")
@admin_required
def get_session_stats():
    db = get_db()
    cursor = db.cursor()
    
    # Get active users (users with ongoing sessions)
    cursor.execute("""
        SELECT COUNT(DISTINCT user_id) as active_users 
        FROM sessions 
        WHERE end_time IS NULL
    """)
    active_users = cursor.fetchone()['active_users']
    
    # Get total sessions
    cursor.execute("SELECT COUNT(*) as total FROM sessions")
    total_sessions = cursor.fetchone()['total']
    
    # Get completed sessions
    cursor.execute("""
        SELECT COUNT(*) as completed 
        FROM sessions 
        WHERE end_time IS NOT NULL
    """)
    completed_sessions = cursor.fetchone()['completed']
    
    # Get average duration of completed sessions
    cursor.execute("""
        SELECT AVG(CAST((julianday(end_time) - julianday(start_time)) * 24 * 60 AS INTEGER)) as avg_duration 
        FROM sessions 
        WHERE end_time IS NOT NULL
    """)
    avg_duration = cursor.fetchone()['avg_duration'] or 0
    
    return jsonify({
        'success': True,
        'active_users': active_users,
        'total_sessions': total_sessions,
        'completed_sessions': completed_sessions,
        'average_duration': avg_duration
    })

@app.route("/api/get_filtered_sessions")
@admin_required
def get_filtered_sessions():
    filter_type = request.args.get('filter', '')
    db = get_db()
    cursor = db.cursor()
    
    query = """
        SELECT 
            s.id, 
            u.username as student_id, 
            u.firstname, 
            u.lastname,
            datetime(s.time_in, 'localtime') as start_time, 
            datetime(s.time_out, 'localtime') as end_time, 
            s.purpose, 
            s.lab_unit,
            s.duration,
            u.remaining_sessions,
            CASE 
                WHEN s.time_out IS NULL THEN 'active' 
                ELSE 'completed' 
            END as status,
            CASE 
                WHEN s.time_out IS NULL THEN 
                    CAST((julianday('now') - julianday(s.time_in)) * 24 * 60 AS INTEGER)
                ELSE 
                    s.duration 
            END as duration
        FROM sit_in_history s
        JOIN users u ON s.user_id = u.id
    """
    
    if filter_type == 'today':
        query += " WHERE DATE(s.time_in) = DATE('now')"
    elif filter_type == 'week':
        query += " WHERE s.time_in >= datetime('now', '-7 days')"
    elif filter_type == 'month':
        query += " WHERE s.time_in >= datetime('now', '-30 days')"
    elif filter_type == 'active':
        query += " WHERE s.time_out IS NULL"
    
    query += " ORDER BY s.time_in DESC"
    
    cursor.execute(query)
    sessions = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
    
    # Log fetched data for debugging
    print(f"Fetched {len(sessions)} sit-in sessions from database")
    if sessions:
        print(f"First session data: {sessions[0]}")
    
    return jsonify({
        'success': True,
        'sessions': sessions
    })

@app.route("/api/export_sessions")
@admin_required
def export_sessions():
    format_type = request.args.get('format', 'csv')  # Default to CSV if not specified
    
    try:
        # Connect to the database 
        conn = sqlite3.connect("users.db")
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # First, let's check which tables exist to determine the right query
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = [table['name'] for table in cursor.fetchall()]
        print(f"Available tables: {tables}")
        
        # Determine if we should use sit_in_history or sessions table
        if 'sit_in_history' in tables:
            # Query from sit_in_history table
            cursor.execute("""
                SELECT s.id, u.username as student_id, u.firstname, u.lastname,
                s.time_in as start_time, s.time_out as end_time, 
                s.purpose, s.lab_unit, u.remaining_sessions,
                s.duration, s.status
                FROM sit_in_history s
                JOIN users u ON s.user_id = u.id
                ORDER BY s.time_in DESC
            """)
        else:
            # Fallback to sessions table
            cursor.execute("""
                SELECT s.id, u.username as student_id, u.firstname, u.lastname,
                s.start_time, s.end_time, 
                '' as purpose, '' as lab_unit, u.remaining_sessions,
                (CASE WHEN s.duration IS NULL AND s.status = 'active' 
                      THEN CAST((julianday(datetime('now')) - julianday(s.start_time)) * 24 * 60 as INTEGER)
                      ELSE s.duration END) as duration,
                s.status
                FROM sessions s
                JOIN users u ON s.user_id = u.id
                ORDER BY s.start_time DESC
            """)
        
        sessions = cursor.fetchall()
        print(f"Found {len(sessions)} records for export")
        
        if format_type == 'csv':
            # Create CSV content
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(['Session ID', 'Student ID', 'Name', 'Start Time', 'End Time', 'Duration (minutes)', 
                            'Purpose', 'Lab Unit', 'Status', 'Remaining Sessions'])
            
            for session in sessions:
                writer.writerow([
                    session['id'],
                    session['student_id'],
                    f"{session['firstname']} {session['lastname']}",
                    session['start_time'],
                    session['end_time'] or 'Ongoing',
                    session['duration'],
                    session['purpose'] or '-',
                    session['lab_unit'] or '-',
                    session['status'],
                    session['remaining_sessions']
                ])
            
            # Create the response
            response = Response(
                output.getvalue(),
                mimetype='text/csv',
                headers={'Content-Disposition': 'attachment; filename=sessions_export.csv'}
            )
            
        elif format_type == 'excel':
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sessions"
            
            # Add header row with formatting
            headers = ['Session ID', 'Student ID', 'Name', 'Start Time', 'End Time', 'Duration (min)', 
                      'Purpose', 'Lab Unit', 'Status', 'Remaining Sessions']
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
            
            # Add data rows
            for row_num, session in enumerate(sessions, 2):
                ws.cell(row=row_num, column=1, value=session['id'])
                ws.cell(row=row_num, column=2, value=session['student_id'])
                ws.cell(row=row_num, column=3, value=f"{session['firstname']} {session['lastname']}")
                ws.cell(row=row_num, column=4, value=session['start_time'])
                ws.cell(row=row_num, column=5, value=session['end_time'] or 'Ongoing')
                ws.cell(row=row_num, column=6, value=session['duration'])
                ws.cell(row=row_num, column=7, value=session['purpose'] or '-')
                ws.cell(row=row_num, column=8, value=session['lab_unit'] or '-')
                ws.cell(row=row_num, column=9, value=session['status'])
                ws.cell(row=row_num, column=10, value=session['remaining_sessions'])
                
                # Highlight active sessions
                if session['status'] == 'active':
                    for col in range(1, 11):
                        ws.cell(row=row_num, column=col).fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Save to a temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            temp_file.close()
            
            # Send file as response
            response = send_file(
                temp_file.name,
                as_attachment=True,
                download_name='sessions_export.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        elif format_type == 'pdf':
            # Create a PDF file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            
            # Create the PDF document
            doc = SimpleDocTemplate(
                temp_file.name,
                pagesize=letter
            )
            
            # Container for the 'Flowable' objects
            elements = []
            
            # Define styles
            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            
            # Add title
            title = Paragraph("CCS Sit-In Sessions Report", title_style)
            elements.append(title)
            elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
            elements.append(Paragraph(" ", styles['Normal']))  # Add some space
            
            # Prepare data for table
            data = [['Session ID', 'Student ID', 'Name', 'Start Time', 'End Time', 'Duration', 'Status', 'Remaining']]
            
            for session in sessions:
                duration_formatted = f"{session['duration']} min"
                
                data.append([
                    session['id'],
                    session['student_id'],
                    f"{session['firstname']} {session['lastname']}",
                    session['start_time'],
                    session['end_time'] or 'Ongoing',
                    duration_formatted,
                    session['status'],
                    session['remaining_sessions']
                ])
            
            # Create the table
            t = Table(data)
            
            # Add style
            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.12, 0.24, 0.45)),  # Header background
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Header text color
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align all cells
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold font for header
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Bottom padding for header
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Table background
                ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Add grid to all cells
            ])
            
            # Add alternating row colors
            for i in range(1, len(data)):
                if i % 2 == 0:
                    bc = colors.Color(0.95, 0.95, 0.95)  # Light gray for even rows
                    style.add('BACKGROUND', (0, i), (-1, i), bc)
            
            t.setStyle(style)
            
            # Add the table to the elements
            elements.append(t)
            
            # Build the PDF
            doc.build(elements)
            temp_file.close()
            
            # Send file as response
            response = send_file(
                temp_file.name,
                as_attachment=True,
                download_name='sessions_export.pdf',
                mimetype='application/pdf'
            )
        
        else:
            # Unsupported format
            conn.close()
            return jsonify({"error": "Unsupported export format"}), 400
        
        # Close database connection
        conn.close()
        return response
        
    except Exception as e:
        # Log the error
        print(f"Error generating report: {str(e)}")
        return jsonify({"error": f"Failed to generate report: {str(e)}"}), 500

@app.route("/admin/announcements")
@admin_required
def admin_announcements():
    return render_template("admin_announcements.html")

@app.route("/api/create_announcement", methods=["POST"])
@admin_required
def create_announcement():
    try:
        data = request.get_json()
        title = data.get("title")
        content = data.get("content")
        priority = data.get("priority", "normal")
        
        if not title or not content:
            return jsonify({"success": False, "error": "Title and content are required"})
        
        if priority not in ["normal", "high", "urgent"]:
            priority = "normal"
        
        conn = sqlite3.connect("users.db")
        cur = conn.cursor()
        
        cur.execute("""
            INSERT INTO announcements (title, content, priority, created_by)
            VALUES (?, ?, ?, ?)
        """, (title, content, priority, session["user"]["id"]))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/get_announcements")
@admin_required
def get_announcements():
    try:
        conn = sqlite3.connect("users.db")
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        cur.execute("""
            SELECT a.*, u.username as created_by
            FROM announcements a
            JOIN users u ON a.created_by = u.id
            ORDER BY a.created_at DESC
        """)
        
        announcements = [dict(row) for row in cur.fetchall()]
        conn.close()
        
        return jsonify({"success": True, "announcements": announcements})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/api/delete_announcement/<int:announcement_id>", methods=["POST"])
@admin_required
def delete_announcement(announcement_id):
    try:
        conn = sqlite3.connect("users.db")
        cur = conn.cursor()
        
        cur.execute("DELETE FROM announcements WHERE id = ?", (announcement_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route("/admin/sitin")
@admin_required
def admin_sitin():
    # Debug session information
    print("Session info:", session)
    print("Is admin in session?", "admin" in session)
    if "admin" in session:
        print("Admin value:", session["admin"])
    print("User info:", session.get("user", {}))
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # Get statistics
        cursor.execute("""
            SELECT 
                COUNT(CASE WHEN time_out IS NULL THEN 1 END) as active_sessions,
                COUNT(*) as total_records,
                COUNT(CASE WHEN DATE(time_in) = DATE('now', 'localtime') THEN 1 END) as today_sessions,
                SUM(CASE 
                    WHEN DATE(time_in) = DATE('now', 'localtime') THEN 
                        CASE 
                            WHEN time_out IS NOT NULL THEN 
                                (julianday(time_out) - julianday(time_in)) * 24
                            ELSE 
                                (julianday('now', 'localtime') - julianday(time_in)) * 24
                        END
                    ELSE 0 
                END) as total_hours_today
            FROM sit_in_history
        """)
        stats_data = cursor.fetchone()
        
        # Format the stats data into a dictionary
        stats = {}
        if stats_data:
            stats = dict(zip(['active_sessions', 'total_records', 'today_sessions', 'total_hours_today'], stats_data))
            
            # Format the total hours
            if stats['total_hours_today'] is not None:
                stats['total_hours_today'] = f"{stats['total_hours_today']:.1f}"
            else:
                stats['total_hours_today'] = "0.0"
        else:
            # Default values if no data found
            stats = {
                'active_sessions': 0,
                'total_records': 0,
                'today_sessions': 0,
                'total_hours_today': "0.0"
            }
        
        # Get sit-in records
        cursor.execute("""
            SELECT 
                s.id as session_id,
                u.username as student_id,
                u.firstname || ' ' || u.lastname as student_name,
                u.course,
                u.yearlevel,
                datetime(s.time_in, 'localtime') as time_in,
                datetime(s.time_out, 'localtime') as time_out,
                CASE
                    WHEN s.time_out IS NOT NULL THEN 
                        ROUND((julianday(s.time_out) - julianday(s.time_in)) * 24 * 60) || ' min'
                    ELSE NULL
                END as duration,
                CASE WHEN s.time_out IS NULL THEN 'active' ELSE 'completed' END as status
            FROM sit_in_history s
            JOIN users u ON s.user_id = u.id
            ORDER BY s.time_in DESC
        """)
        records = [dict(zip([column[0] for column in cursor.description], row)) for row in cursor.fetchall()]
        
        cursor.close()
        
        return render_template("admin_sitin.html", stats=stats, records=records)
    except Exception as e:
        print(f"Error in admin_sitin: {str(e)}")
        flash('Error: ' + str(e), 'error')
        return redirect(url_for('admin_dashboard'))

@app.route("/api/export_sitin_records")
@admin_required
def export_sitin_records():
    conn = get_db()
    cur = conn.cursor()

    # Get filter parameters
    date = request.args.get('date')
    status = request.args.get('status')
    search = request.args.get('search')

    # Base query
    query = """
        SELECT 
            u.username as student_id,
            u.firstname || ' ' || u.lastname as name,
            u.course,
            u.yearlevel,
            datetime(s.time_in) as time_in,
            datetime(s.time_out) as time_out,
            CASE 
                WHEN s.time_out IS NOT NULL 
                THEN round((julianday(s.time_out) - julianday(s.time_in)) * 24, 1)
                ELSE NULL 
            END as duration,
            s.status
        FROM sit_in_history s
        JOIN users u ON s.user_id = u.id
        WHERE 1=1
    """
    params = []

    # Add filters
    if date:
        query += " AND date(s.time_in) = ?"
        params.append(date)
    if status:
        query += " AND s.status = ?"
        params.append(status)
    if search:
        query += " AND (u.username LIKE ? OR u.firstname LIKE ? OR u.lastname LIKE ?)"
        search_param = f"%{search}%"
        params.extend([search_param, search_param, search_param])

    query += " ORDER BY s.time_in DESC"
    
    cur.execute(query, params)
    records = cur.fetchall()

    # Create CSV file
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Student ID', 'Name', 'Course', 'Year Level', 'Time In', 'Time Out', 'Duration (hours)', 'Status'])
    
    for record in records:
        writer.writerow([
            record[0],  # student_id
            record[1],  # name
            record[2],  # course
            record[3],  # year level
            record[4],  # time_in
            record[5],  # time_out
            record[6],  # duration
            record[7]   # status
        ])

    # Create response
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=sitin_records.csv'}
    )

@app.route('/submit_feedback', methods=['POST'])
def submit_feedback():
    if "user" not in session:
        return jsonify({"error": "Please log in first."}), 401
    
    data = request.get_json()
    user_id = session["user"]["id"]
    current_time = datetime.now()
    
    try:
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        
        # Insert feedback
        cur.execute("""
            INSERT INTO feedback (user_id, rating, comments, created_at)
            VALUES (?, ?, ?, ?)
        """, (user_id, data['rating'], data['comments'], current_time))
        
        con.commit()
        con.close()
        
        return jsonify({"success": True, "message": "Feedback submitted successfully"})
    except Exception as e:
        if con:
            con.close()
        return jsonify({"error": str(e)}), 500

@app.route("/api/get_feedback")
@admin_required
def get_feedback():
    try:
        con = sqlite3.connect("users.db")
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        
        # Get feedback with user details
        cur.execute("""
            SELECT 
                f.id,
                f.rating,
                f.comments,
                f.created_at,
                u.username,
                u.firstname,
                u.lastname
            FROM feedback f
            JOIN users u ON f.user_id = u.id
            ORDER BY f.created_at DESC
        """)
        feedback = [dict(row) for row in cur.fetchall()]
        
        # Get statistics
        cur.execute("""
            SELECT 
                ROUND(AVG(rating), 2) as avg_rating,
                COUNT(*) as total,
                SUM(CASE WHEN rating >= 4 THEN 1 ELSE 0 END) as positive,
                SUM(CASE WHEN created_at >= datetime('now', '-7 days') THEN 1 ELSE 0 END) as recent
            FROM feedback
        """)
        stats = dict(cur.fetchone())
        
        con.close()
        
        return jsonify({
            'success': True,
            'feedback': feedback,
            'stats': {
                'average_rating': float(stats['avg_rating'] or 0),
                'total_feedback': stats['total'],
                'positive_ratings': stats['positive'],
                'recent_feedback': stats['recent']
            }
        })
    except Exception as e:
        if con:
            con.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route("/api/export_feedback")
@admin_required
def export_feedback():
    try:
        con = sqlite3.connect("users.db")
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        
        cur.execute("""
            SELECT 
                f.created_at,
                f.rating,
                f.comments,
                u.username,
                u.firstname,
                u.lastname
            FROM feedback f
            JOIN users u ON f.user_id = u.id
            ORDER BY f.created_at DESC
        """)
        
        feedback_data = cur.fetchall()
        con.close()
        
        # Create CSV content
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Date', 'Student ID', 'Name', 'Rating', 'Comments'])
        
        for feedback in feedback_data:
            writer.writerow([
                feedback['created_at'],
                feedback['username'],
                f"{feedback['firstname']} {feedback['lastname']}",
                feedback['rating'],
                feedback['comments']
            ])
        
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=feedback_export.csv'}
        )
    except Exception as e:
        if con:
            con.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route("/admin/feedback")
@admin_required
def admin_feedback():
    return render_template("admin_feedback.html")

@app.route("/api/add_user", methods=["POST"])
@admin_required
def add_user():
    try:
        data = request.json
        
        # Generate a random password for the new user
        password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
        hashed_password = generate_password_hash(password, method="pbkdf2:sha256")
        
        # Generate a username (student ID) based on timestamp
        username = f"ST{datetime.now().strftime('%y%m%d%H%M%S')}"
        
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        
        # Insert new user
        cur.execute("""
            INSERT INTO users (firstname, lastname, email, username, password, course, yearlevel, profile_pic, remaining_sessions)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 30)
        """, (data['firstname'], data['lastname'], data['email'], username, hashed_password, 
              data['course'], data['yearlevel'], "default.png"))
        
        user_id = cur.lastrowid
        con.commit()
        
        return jsonify({
            "success": True,
            "message": "User added successfully",
            "user": {
                "id": user_id,
                "username": username,
                "password": password  # Send the plain password back so admin can share it with the user
            }
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        if 'con' in locals():
            con.close()

@app.route("/api/student_info", methods=["POST"])
@admin_required
def get_student_info():
    data = request.json
    student_id = data.get('student_id')
    
    if not student_id:
        return jsonify({"success": False, "error": "No student ID provided"})
    
    try:
        con = sqlite3.connect("users.db")
        cur = con.cursor()
        
        # Get student details including remaining_sessions
        cur.execute("""
            SELECT id, firstname, lastname, course, yearlevel, remaining_sessions
            FROM users
            WHERE username = ? AND is_admin = 0
        """, (student_id,))
        
        student = cur.fetchone()
        
        if not student:
            return jsonify({"success": False, "error": "Student not found"})
        
        user_id = student[0]
        
        # Check if student has active sessions
        cur.execute("""
            SELECT 
                CASE 
                    WHEN EXISTS (SELECT 1 FROM sessions WHERE user_id = ? AND status = 'active')
                    OR EXISTS (SELECT 1 FROM sit_in_history WHERE user_id = ? AND status = 'active')
                    THEN 1
                    ELSE 0
                END as has_active_session
        """, (user_id, user_id))
        
        has_active_session = cur.fetchone()[0]
        status = "Active Session" if has_active_session else "Available"
        
        student_info = {
            "success": True,
            "student_name": f"{student[1]} {student[2]}",
            "course": student[3],
            "yearlevel": student[4],
            "status": status,
            "user_id": user_id,
            "remaining_sessions": student[5]
        }
        
        return jsonify(student_info)
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        if 'con' in locals() and con:
            con.close()

@app.route("/api/update_remaining_sessions", methods=["POST"])
@admin_required
def update_remaining_sessions():
    try:
        data = request.json
        student_id = data.get('student_id')
        new_count = data.get('remaining_sessions')
        
        if not student_id or new_count is None:
            return jsonify({"success": False, "error": "Student ID and remaining sessions count are required"})
        
        # Ensure new_count is a non-negative integer
        try:
            new_count = int(new_count)
            if new_count < 0:
                return jsonify({"success": False, "error": "Remaining sessions count cannot be negative"})
        except ValueError:
            return jsonify({"success": False, "error": "Remaining sessions count must be a valid number"})
        
        conn = sqlite3.connect("users.db")
        cur = conn.cursor()
        
        # Check if student exists
        cur.execute("SELECT id FROM users WHERE username = ?", (student_id,))
        if not cur.fetchone():
            conn.close()
            return jsonify({"success": False, "error": "Student not found"})
        
        # Update remaining sessions
        cur.execute("""
            UPDATE users 
            SET remaining_sessions = ?
            WHERE username = ?
        """, (new_count, student_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "success": True,
            "message": f"Updated remaining sessions to {new_count} for student {student_id}"
        })
    except Exception as e:
        if 'conn' in locals():
            conn.close()
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/get_purpose_stats')
@admin_required
def get_purpose_stats():
    try:
        conn = sqlite3.connect('users.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get sit-in purpose distribution for the chart
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN purpose IS NULL OR purpose = '' THEN 'Other'
                    ELSE purpose 
                END as purpose, 
                COUNT(*) as count 
            FROM sit_in_history 
            GROUP BY purpose
            ORDER BY count DESC
        """)
        purpose_stats = [dict(row) for row in cursor.fetchall()]
        
        return jsonify({
            "success": True, 
            "purpose_stats": purpose_stats
        })
        
    except Exception as e:
        return jsonify({
            "success": False, 
            "error": str(e)
        })
    finally:
        if 'conn' in locals():
            conn.close()

# Route to manage lab resources
@app.route("/admin/lab_resources")
@admin_required
def admin_lab_resources():
    conn = sqlite3.connect('users.db')
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT * FROM lab_resources")
    resources = [dict(row) for row in cur.fetchall()]
    conn.close()
    return render_template("admin_lab_resources.html", resources=resources)

@app.route("/api/toggle_resource/<int:resource_id>", methods=["POST"])
@admin_required
def toggle_resource(resource_id):
    try:
        conn = sqlite3.connect("users.db")
        cur = conn.cursor()
        cur.execute("SELECT status FROM lab_resources WHERE id = ?", (resource_id,))
        resource = cur.fetchone()
        if not resource:
            return jsonify({"success": False, "error": "Resource not found"})
        new_status = "disabled" if resource[0] == "enabled" else "enabled"
        cur.execute("UPDATE lab_resources SET status = ? WHERE id = ?", (new_status, resource_id))
        conn.commit()
        return jsonify({"success": True, "new_status": new_status})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        conn.close()

# Route to manage lab usage points
@app.route("/admin/lab_points")
@admin_required
def admin_lab_points():
    conn = sqlite3.connect('users.db')
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT id, username, points FROM users WHERE is_admin = 0")
    users = [dict(row) for row in cur.fetchall()]
    conn.close()
    return render_template("admin_lab_points.html", users=users)

@app.route("/api/update_points/<int:user_id>", methods=["POST"])
@admin_required
def update_points(user_id):
    data = request.json
    points = data.get('points')
    try:
        conn = sqlite3.connect("users.db")
        cur = conn.cursor()
        cur.execute("UPDATE users SET points = ? WHERE id = ?", (points, user_id))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})
    finally:
        conn.close()

# Route to display leaderboard
@app.route("/admin/leaderboard")
@admin_required
def admin_leaderboard():
    conn = sqlite3.connect('users.db')
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("SELECT username, points FROM users WHERE is_admin = 0 ORDER BY points DESC LIMIT 10")
    leaderboard = [dict(row) for row in cur.fetchall()]
    conn.close()
    return render_template("admin_leaderboard.html", leaderboard=leaderboard)

# Route to generate reports
@app.route("/admin/reports")
@admin_required
def admin_reports():
    return render_template("admin_reports.html")

@app.route("/api/generate_report", methods=["POST"])
@admin_required
def generate_report():
    data = request.json
    report_type = data.get('report_type')
    # Implement report generation logic here
    return jsonify({"success": True, "message": "Report generated successfully"})

if __name__ == "__main__":
    # Initialize the database before starting the app
    init_db()
    app.run(debug=True, port="5000")
